import openpyxl
book = openpyxl.load_workbook('C:\\Users\\MOUNIKA\\Desktop\\python.xlsx')
sheet = book.active
cell = sheet.cell(row=1, column=1)
print(cell.value)
sheet.cell(row=2, column=3).value = "AsulaMounika"
print(sheet.cell(row=2, column=3).value)
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row = i, column = 1).value == "Mounika":
        for j in range(2,sheet.max_column+1):

            print(sheet.cell(row=i , column=j).value)